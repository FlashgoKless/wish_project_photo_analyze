import face_recognition as fr
from os import listdir
import openpyxl
#import cv2

list_name_img = list(filter(lambda x: x.endswith('.jpg'), listdir())) #чтение файлов .jpg

#сортировка списка имён файлов .jpg
for i in range(len(list_name_img)):
    for j in range(i+1, len(list_name_img)):
        if int(list_name_img[i][0:-4]) > int(list_name_img[j][0:-4]):
            list_name_img[i],list_name_img[j] = list_name_img[j],list_name_img[i]


list_lovations_for_cv = []
list_encodings = []
col = 1

xl_file = openpyxl.load_workbook('encodings.xlsx') #загрузка xlsx файла
sheet = xl_file.active #активация листа

for name in list_name_img:
    sheet.cell(row = 1, column = col).value = name
    image_fr = fr.load_image_file(str(name)) #загрузка изображения
    list_lovations_for_cv.append(fr.face_locations(image_fr)) #для прямоугольников по контуру лиц
    list_encodings = fr.face_encodings(image_fr)
    for row in range(128):
        sheet.cell(row=row+2, column=col).value = list_encodings[0][row]
    col += 1

xl_file.save('encodings.xlsx')



'''
###face in rectangle###

top = face_locations[0][0]
right = face_locations[0][1]
bottom = face_locations[0][2]
left = face_locations[0][3]
image_cv = cv2.imread('novak.jpg')
output = cv2.rectangle(image_cv, (left, top), (right, bottom), (255, 0, 0), 2)
cv2.imshow('kkk', output)
cv2.waitKey(0)
'''
